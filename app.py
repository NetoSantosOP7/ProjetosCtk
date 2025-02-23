import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl, xlrd
from openpyxl import load_workbook
import pathlib
from openpyxl import Workbook

#Setando a aparencia padrao do sistema

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearence()
        self.todo_sistema()
    
    def layout_config(self):
        self.title("Sistema de Gestão de Clientes")
        self.geometry("700x500")

    def appearence(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent", text_color=['#000', '#fff']).place(x=50, y=430)
        self.opt_apm = ctk.CTkOptionMenu(self, values=["Light", "Dark", "System"], command=self.change_apm).place(x=50, y=460)

    def todo_sistema(self):
        frame = ctk.CTkFrame(self, width=700, height=50, corner_radius=0, bg_color="teal", fg_color="teal").place(x=0, y=10)
        title = ctk.CTkLabel(frame, text="Sistema de Gestão de Clientes", font=("Century Gothic bold", 24), text_color="#fff", bg_color="teal").place(x=190, y=10)

        span = ctk.CTkLabel(self, text="Por favor, preencha todos os campos do formulário!", font=("Century Gothic bold", 16), text_color=["#000", "#fff"]).place(x=50, y=70)

        ficheiro = pathlib.Path("Clientes.xlsx")

        if ficheiro.exists():
            pass
        else:
            ficheiro = Workbook()
            folha = ficheiro.active
            folha['A1'] = "Nome Completo"
            folha['B1'] = "Contato"
            folha['C1'] = "Idade"
            folha['D1'] = "Gênero"
            folha['E1'] = "Endereço"
            folha['F1'] = "Observações"

            ficheiro.save("Clientes.xlsx")

        def submit():

            #Pegando os dados dos entrys
            name = name_value.get()
            contact = contact_entry.get()
            age = age_value.get()
            gender = gender_combobox.get()
            obs = obs_entry.get(0.0, END)
            adress = address_value.get()

            if (name=="" or contact=="" or age=="" or gender=="" or obs=="" or adress==""):
                messagebox.showwarning("Sistema", "ERRO!\nPor favor, preencha todos os campos do formulário!")

            ficheiro = openpyxl.load_workbook("Clientes.xlsx")

            folha = ficheiro.active

            folha.cell(column=1, row=folha.max_row+1, value=name)
            folha.cell(column=2, row=folha.max_row, value=contact)
            folha.cell(column=3, row=folha.max_row, value=age)
            folha.cell(column=4, row=folha.max_row, value=gender)
            folha.cell(column=5, row=folha.max_row, value=address_value.get())
            folha.cell(column=6, row=folha.max_row, value=obs)

            ficheiro.save("Clientes.xlsx")
            messagebox.showinfo("Sistema", "Os dados foram salvos com sucesso!")

        def clear():
            name_value.set("")
            contact_value.set("")
            age_value.set("")
            obs_entry.delete("0.0", END)
            address_value.set("")


        #Texts variables
        name_value = StringVar()
        contact_value = StringVar()
        age_value = StringVar()
        address_value = StringVar()
        obs_value = StringVar()

        #Entrys
        name_entry = ctk.CTkEntry(self, width=350,textvariable=name_value, font=("Century  bold", 16), placeholder_text="Nome Completo", fg_color="transparent")
        contact_entry = ctk.CTkEntry(self, width=200, textvariable=contact_value, font=("Century  bold", 16), placeholder_text="Número ou Email de contato", fg_color="transparent")
        age_entry = ctk.CTkEntry(self, width=150, textvariable=age_value, font=("Century  bold", 16), placeholder_text="Digite sua idade", fg_color="transparent")
        address_entry = ctk.CTkEntry(self, width=200,textvariable=address_value, font=("Century  bold", 16), placeholder_text="Digite seu endereço", fg_color="transparent")

        #Combobox
        gender_combobox = ctk.CTkComboBox(self, values=["Masculino", "Feminino"], font=("Century Gothic bold", 14), width=150)
        gender_combobox.set("Masculino")

        #Entrada de Observações
        obs_entry = ctk.CTkTextbox(self, width=450, height=150, font=("Arial", 18), fg_color="transparent", border_color="#aaa", border_width=2)

        #Labels
        lb_name = ctk.CTkLabel(self, text="Nome Completo", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_contact = ctk.CTkLabel(self, text="Contato", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_age = ctk.CTkLabel(self, text="Idade", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_gander = ctk.CTkLabel(self, text="Gênero", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_address = ctk.CTkLabel(self, text="Endereço", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_obs = ctk.CTkLabel(self, text="Observações", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])

        btn_submit = ctk.CTkButton(self, text="Salvar dados".upper(), command=submit, fg_color="#151", hover_color="#131").place(x=300, y=420)
        btn_submit = ctk.CTkButton(self, text="Limpar campos".upper(), command=clear, fg_color="#151", hover_color="#131").place(x=500, y=420)
        
        #Posicionando os elementos na janela

        lb_name.place(x=50, y=120)
        name_entry.place(x=50, y=150)

        lb_contact.place(x=450, y=120)
        contact_entry.place(x=450, y=150)

        lb_age.place(x=300, y=190)
        age_entry.place(x=300, y=220)

        lb_gander.place(x=500, y=190)
        gender_combobox.place(x=500, y=220)

        lb_address.place(x=50, y=190)
        address_entry.place(x=50, y=220)

        lb_obs.place(x=50, y=260)
        obs_entry.place(x=180, y=260)

    def change_apm(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)

if __name__ =="__main__":
    app = App()
    app.mainloop()

